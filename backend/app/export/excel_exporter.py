"""
Excel Exporter - generates a downloadable Excel timetable.
"""
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from app.schemas.faculty import FacultyTimetableResponse


class ExcelExporter:
    """Generates Excel files for faculty timetables."""

    @staticmethod
    def export(timetable: FacultyTimetableResponse) -> BytesIO:
        """Generate Excel and return as BytesIO buffer."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Timetable"

    @staticmethod
    def _fill_sheet(ws, timetable: FacultyTimetableResponse):
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        import openpyxl

        # Styles
        header_font = Font(bold=True, size=14)
        col_header_font = Font(bold=True)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        break_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")

        # Title
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(timetable.headers) + 1)
        title_cell = ws.cell(row=1, column=1, value=f"Timetable: {timetable.faculty_name}")
        title_cell.font = header_font
        title_cell.alignment = center_align

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(timetable.headers) + 1)
        sub_cell = ws.cell(row=2, column=1, value="Department: ")
        sub_cell.alignment = center_align

        # Headers
        ws.cell(row=4, column=1, value="Day / Period").font = col_header_font
        ws.cell(row=4, column=1).border = thin_border
        
        for col_idx, header in enumerate(timetable.headers, start=2):
            cell = ws.cell(row=4, column=col_idx, value=str(header))
            cell.font = col_header_font
            cell.alignment = center_align
            cell.border = thin_border
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20

        ws.column_dimensions['A'].width = 15

        # Body
        start_row = 5
        for i, day in enumerate(timetable.days):
            row_idx = start_row + i
            
            # Day column
            day_cell = ws.cell(row=row_idx, column=1, value=str(day))
            day_cell.font = col_header_font
            day_cell.border = thin_border
            day_cell.alignment = center_align

            # Period columns
            for j, cell_data in enumerate(timetable.schedule[i], start=2):
                cell = ws.cell(row=row_idx, column=j)
                cell.border = thin_border
                cell.alignment = center_align

                if cell_data.type == "class":
                    cell.value = f"{cell_data.subject_code}\n{cell_data.class_name}\n{cell_data.room}"
                elif cell_data.type == "break":
                    cell.value = "BREAK"
                    cell.fill = break_fill

    @staticmethod
    def export(timetable: FacultyTimetableResponse) -> BytesIO:
        """Generate Excel and return as BytesIO buffer."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Timetable"
        
        ExcelExporter._fill_sheet(ws, timetable)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def export_all(timetables: list[FacultyTimetableResponse]) -> BytesIO:
        """Generate a multi-sheet Excel for all faculties."""
        import re
        wb = openpyxl.Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        for timetable in timetables:
            # Excel sheet names can't contain certain characters and must be <= 31 chars
            safe_name = re.sub(r'[\\*?:/\[\]]', '', timetable.faculty_name)[:31]
            # Ensure unique sheet names
            base_name = safe_name
            counter = 1
            while safe_name in wb.sheetnames:
                suffix = str(counter)
                safe_name = base_name[:31 - len(suffix)] + suffix
                counter += 1
                
            ws = wb.create_sheet(title=safe_name)
            ExcelExporter._fill_sheet(ws, timetable)
            
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
